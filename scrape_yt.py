import os
import streamlit as st
from playwright.sync_api import sync_playwright
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime
import re
import io

os.system("playwright install chromium")

def parse_youtube_number(text):
    text = text.upper().strip()
    if 'K' in text:
        return int(float(text.replace('K', '')) * 1000)
    if 'M' in text:
        return int(float(text.replace('M', '')) * 1000000)
    if text.isdigit():
        return int(text)
    return text

def scrape(page, url: str):
    page.goto(url)

    try:
        page.locator('button', has_text='Reject all').click(timeout=3000)
    except:
        pass

    try:
        is_live = page.locator('meta[itemprop="isLiveBroadcast"]').count() > 0
    except:
        is_live = "need to see it manual"

    page.evaluate("window.scrollBy(0, 600)")

    try:
        pattern = re.compile(r"like this video|me gusta", re.IGNORECASE)
        likes_locator = page.get_by_role("button", name=pattern).first
        likes_locator.wait_for(timeout=5000)
        
        aria_text = likes_locator.get_attribute("aria-label") or ""
        numeric_likes_string = "".join(filter(str.isdigit, aria_text))
        likes_count = int(numeric_likes_string) if numeric_likes_string else "need to see it manual"
    except:
        likes_count = "need to see it manual"
        

    try:
        views_locator = page.locator('#info span.style-scope.yt-formatted-string').first
        views_locator.wait_for(timeout=5000)
        raw_views = views_locator.inner_text().split(' ')[0]
        clean_views = raw_views.strip()
        views_count = parse_youtube_number(clean_views)
    except:
        views_count = "need to see it manual"

    try:
        commentSelector = 'ytd-comments-header-renderer #count yt-formatted-string span'
        page.wait_for_selector(commentSelector, timeout=5000)
        comments_text = page.locator(commentSelector).first.inner_text()
        
        if '\n' in comments_text:
            raise ValueError("Animated comment count detected")
            
        clean_comments = comments_text.replace(',', '').strip()
        if not clean_comments:
            raise ValueError("Empty string")
    except:
        clean_comments = "need to see it manual"

    return views_count, likes_count, clean_comments, is_live

def read_urls_from_excel(uploaded_file):
    workbook = openpyxl.load_workbook(uploaded_file)
    sheet = workbook.worksheets[0] 
    urls = []
    
    for row in sheet.iter_rows(min_row=2, max_col=1, values_only=True):
        if row[0]:
            urls.append(str(row[0]).strip())
            
    return urls

def generate_excel_in_memory(data):
    wb = openpyxl.Workbook()
    ws = wb.worksheets[0] 
    ws.title = "Scrape Results"
    
    headers = ['KOL Type', 'KOL', 'Date', 'Week', 'Month', 'Platform', 'Link', 'Game', 'Views', 'Comments', 'Likes']
    ws.append(headers)
    
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        
    for row in data:
        ws.append(row)
        
    for col_idx, col in enumerate(ws.columns, 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        
        for cell in col:
            if cell.value is not None:
                try:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
                except:
                    pass
                    
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

st.set_page_config(page_title="App")
st.title("Bro me tienes haciendo tu chamba")

col1, col2, col3 = st.columns(3)
with col1:
    kol_type = st.text_input('KOL Type:')
with col2:
    kol = st.text_input('KOL:')
with col3:
    game = st.text_input('Game:')

uploaded_file = st.file_uploader("Sube tu archivo excel.xlsx", type=["xlsx"])

if st.button("Empezar!!!"):
    if not uploaded_file:
        st.error("Socio/a suba el archivo primero 💀")
    else:
        video_list = read_urls_from_excel(uploaded_file)
        
        if not video_list:
            st.error("Links invalidos")
        else:
            st.info(f"Se encontraron {len(video_list)} links....")
            
            now = datetime.now()
            current_date = now.strftime("%d/%m/%Y")
            current_week = now.isocalendar()[1]
            current_month = now.month
            
            results_data = []
            
            progress_bar = st.progress(0)
            status_text = st.empty()

            with sync_playwright() as p:
                browser = p.chromium.launch(headless=True)
                context = browser.new_context(viewport={'width': 1280, 'height': 720})
                page = context.new_page()
                
                for index, link in enumerate(video_list):
                    progress_bar.progress((index + 1) / len(video_list))                   
                    views, likes, comments, is_live = scrape(page, link)
                    
                    row_data = [
                        kol_type,
                        kol,
                        current_date,
                        current_week,
                        current_month,
                        'Youtube Live' if is_live else 'Youtube',
                        link,
                        game,
                        views,
                        comments,
                        likes
                    ]
                    results_data.append(row_data)
                    
                browser.close()

            status_text.success("Completado")
            
            excel_file = generate_excel_in_memory(results_data)
            
            st.download_button(
                label="Desacargar el archivo",
                data=excel_file,
                file_name="result.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )